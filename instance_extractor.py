# -*- coding: utf-8 -*-
"""
Created on Wed Oct 16 11:39:08 2024

@author: Julius de Clercq
"""

import os
import py7zr
# from glob import glob
import pandas as pd
from openpyxl import load_workbook


#%%

def extract_7z_recursive(file_path, extract_to):
    # Ensure the extraction directory exists
    os.makedirs(extract_to, exist_ok=True)
    # Open the .7z file and extract it
    with py7zr.SevenZipFile(file_path, mode='r') as archive:
        archive.extractall(path=extract_to)
    # print(f"Extraction complete. Files extracted to: {extract_to}")
    for root, _, files in os.walk(extract_to):
        for file in files:
            if file.endswith(".7z"):
                nested_7z_path = os.path.join(root, file)
                nested_extract_to = os.path.join(extract_to, os.path.splitext(file)[0])  # Extract in subfolder with the file name
                extract_7z_recursive(nested_7z_path, nested_extract_to)

#%%

def read_instance_file(file_path):
    # file_path = os.path.join(directory, os.listdir(directory)[0])
    with open(file_path, 'r') as file:
        lines = file.readlines()[1:]  # Ignore the first row
    data = []
    column_names = []
    for line in lines:
        numbers = list(map(int, line.split()))
        # Odd-indexed values are the column names
        for i in range(0, len(numbers), 2):
            if i < len(numbers) - 1:
                # Append the column name if it's not already added
                if numbers[i] not in column_names:
                    column_names.append(numbers[i])
                # Append the corresponding value
                data.append(numbers[i + 1])
    # Reshape the data into a DataFrame format
    num_rows = len(lines)
    num_columns = len(column_names)
    # Create a DataFrame with the extracted data
    df = pd.DataFrame(data)
    df = df.values.reshape(num_rows, num_columns)  # Reshape into the correct form

    # Create the final DataFrame
    return pd.DataFrame(df, columns=column_names)


#%%

def instance_extraction(df):
    instances = {i : None for i in df["Instance name"]}
    for instance in df["Instance name"]:
        instance_found = False
        if not instance_found:
            for size in ["Large", "Small"]:
                directory = f"InstancesAndBounds/VRF_Instances/{size}/"
                for filename in os.listdir(directory):
                    if "VFR" in filename: # All the text files have the typo that "VFR" is written instead of "VRF".
                        os.rename(os.path.join(directory, filename), os.path.join(directory, filename.replace("VFR", "VRF")))
                        filename = filename.replace("VFR", "VRF")
                    if filename.startswith(instance):
                        instances[instance] = read_instance_file(os.path.join(directory, filename))
                        instance_found = True
    return instances

#%%

# def write_instance_sheets(instances, instance_excel_file):
#     existing_sheets = pd.read_excel(instance_excel_file, sheet_name=None)
#     with pd.ExcelWriter(instance_excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
#         # Preserve the first sheet
#         first_sheet_name = list(existing_sheets.keys())[0]
#         first_sheet = existing_sheets[first_sheet_name]
#         first_sheet.to_excel(writer, sheet_name=first_sheet_name, index=False)

#         # Add new instance sheets
#         for instance_name, df in instances.items():
#             if instance_name == first_sheet_name:
#                 print(f"Skipping overwriting the first sheet '{first_sheet_name}'")
#                 continue
#             df.to_excel(writer, sheet_name=instance_name, index=True)
            # worksheet = writer.sheets[instance_name]
            # worksheet.write(0, 0, "jobs/machines")
#             print(f"Added sheet '{instance_name}' to {instance_excel_file}")

def write_instance_sheets(instances, instance_excel_file):
    with pd.ExcelWriter(instance_excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Read the existing sheets from the file
            existing_sheets = pd.read_excel(instance_excel_file, sheet_name=None)

            # Preserve the first sheet
            first_sheet_name = list(existing_sheets.keys())[0]
            first_sheet = existing_sheets[first_sheet_name]
            first_sheet.to_excel(writer, sheet_name=first_sheet_name, index=False)

            # Add new instance sheets
            for instance_name, df in instances.items():
                if instance_name == first_sheet_name:
                    print(f"Skipping overwriting the first sheet '{first_sheet_name}'")
                    continue

                # Write each DataFrame to a separate sheet with the instance name as the sheet name
                df.to_excel(writer, sheet_name=instance_name, index=True)

                print(f"Added sheet '{instance_name}' to {instance_excel_file}")


    workbook = load_workbook(instance_excel_file)
    for instance_name in instances.keys():
        if instance_name == first_sheet_name:
            continue
        # Access the worksheet by name
        worksheet = workbook[instance_name]

        # Write "jobs/machines" to the first cell (A1)
        worksheet.cell(row=1, column=1, value="jobs/machines")

    # Save the workbook after making changes
    workbook.save(instance_excel_file)

#%%

# Uncomment the following lines to extract the instances from the .7z file.
# file_path = 'InstancesAndBounds.7z'
# extract_to = 'InstancesAndBounds'
# extract_7z_recursive(file_path, extract_to)

instance_excel_file = "Instance_test.xlsx"
df = pd.read_excel(instance_excel_file)
instances = instance_extraction(df)
write_instance_sheets(instances, instance_excel_file)


















